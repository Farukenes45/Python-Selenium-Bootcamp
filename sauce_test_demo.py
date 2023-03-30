from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import globalConstants

class Test_DemoClass:

    # her testten önce çağrılır
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        # günün tarihini al bu tarih ile bir klasör var mı kontrol et yoksa oluştur
        self.folderPath = str(date.today())
        #exist_ok_True ilgili path oluşturulmuş ise tekrar bunu oluşturma
        Path(self.folderPath).mkdir(exist_ok=True)
    # her testten sonra çağrılır
    def teardown_method(self):
        self.driver.quit()
    # setup_method -> test_demoFunc -> teardown_method bu sırayla çalışır
    def test_demoFunc(self):
        # 3A Act Arrange Assert
        text = "hello"
        # test classımıza bütün olaylar sonucu buysa True ya da False döner
        # assert: ilgili testin bağlı olduğu condition (şartı) belirtir
        assert text == "hello"
    # setup_method -> test_demoFunc2 -> teardown_method bu sırayla çalışır
    def test_demoFunc2(self):
        assert True

    # deperatörlerden çağıracağımız fonksiyonlara self parametresini vermiyoruz
    # veriyi al
    def getData():
        # excel dosyasını yükle
        excel_file = openpyxl.load_workbook("data/invalid_login.xlsx")
        # çalışmak istediğimiz excel sayfasını belirttik
        selectedSheet = excel_file["Sayfa1"]
        total_rows = selectedSheet.max_row
        data = []
        for i in range(2,total_rows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tuple_data = (username,password)
            data.append(tuple_data)
        return data

    #  bu testi geç 
    # @pytest.mark.skip()
    # parametrize fonksiyonu içerisine verdiğimiz parametreleri ve bunlara karşılık gelen değerleri dener
    
    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):

        self.wait_for_element_visible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")

        self.wait_for_element_visible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")

        usernameInput.send_keys(username)
        passwordInput.send_keys(password)

        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()

        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_invalid_login_{username}_{password}.png")
        
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
       
    def wait_for_element_visible(self,locator,timeout=5):
        WebDriverWait(self.driver,5).until(EC.visibility_of_element_located(locator))







